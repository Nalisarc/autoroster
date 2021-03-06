#!/usr/bin/env python
import pandas
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side
def read_in_roster(path):
    """reads in the roster report, and skips to the headers """
    df = pandas.read_excel(path, skiprows=2)
    return df


def get_exam_list(dataframe):
    raw_exam_list = list(dataframe["SP Exam"].unique())
    exam_list = [exam for exam in raw_exam_list if str(
        type(exam)) == "<class 'str'>"]
    exam_list.sort()
    return exam_list


def get_student_info(exam, dataframe):
    student_info = pandas.DataFrame(
        dataframe, columns=["Student Name", "No Show", "Completed"])
    by_exam = dataframe["SP Exam"] == exam
    processed = student_info[by_exam]
    sorted_info = processed.sort_values("Student Name")
    return sorted_info

def make_workbook():
    return openpyxl.Workbook()

def make_sheet(workbook, exam, date, student_info):
    # Text Constants
    title = "Exam Roster Report"
    headers = ["Student Name", "No Show", "Completed", "Check 1", "Check 2"]
    # Create workbook
    ws = workbook.create_sheet(exam)

    # Column widths
    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    # Sets up people field
    ws["A1"] = title
    ws["B1"] = exam
    ws["A3"] = "Date: " + date
    ws["D3"] = "Exam Count: ________"
    ws["A5"] = "Check #1: ___________"
    ws["C5"] = "Check #2: ___________"

    ws["A8"] = headers[0]
    ws["B8"] = headers[1]
    ws["C8"] = headers[2]
    ws["D8"] = headers[3]
    ws["E8"] = headers[4]

    for r in dataframe_to_rows(student_info, index=False, header=False):
        ws.append(r)

    set_border(ws,"A8:E{0}".format(ws.max_row))

    return None


def save_workbook(workbook,path):
    workbook.save(path)


def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border
